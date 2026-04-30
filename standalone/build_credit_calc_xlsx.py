"""Build Credit Payment Calculator as a single .xlsx file.

Opens natively in Numbers (iPad/iOS), Excel, Google Sheets. Mirrors the
logic of credit_calculator.html — provinces, Flexiti + MHC tiers,
cumulative-item subtotals, down payment, monthly payment formulas, and
"longest eligible plan" summary blocks.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule

PROVINCES = [
    ('BC', 'British Columbia',          0.12),
    ('AB', 'Alberta',                   0.05),
    ('SK', 'Saskatchewan',              0.11),
    ('MB', 'Manitoba',                  0.12),
    ('ON', 'Ontario',                   0.13),
    ('QC', 'Quebec',                    0.14975),
    ('NB', 'New Brunswick',             0.15),
    ('NS', 'Nova Scotia',               0.15),
    ('PE', 'Prince Edward Island',      0.15),
    ('NL', 'Newfoundland and Labrador', 0.15),
    ('NT', 'Northwest Territories',     0.05),
    ('NU', 'Nunavut',                   0.05),
    ('YT', 'Yukon',                     0.05),
]

FLEXITI = [   # months, adminFee, minAmount
    (3,  0,      0),
    (6,  24.99,  500),
    (12, 49.99,  1000),
    (18, 74.99,  1500),
    (24, 99.99,  2000),
    (45, 249.99, 4500),
]

MHC = [       # months, adminFee, annualFee, minAmount
    (6,  0, 25, 500),
    (12, 0, 25, 1000),
    (18, 0, 25, 1500),
    (24, 0, 25, 2000),
]

# Styles
H1 = Font(name='Helvetica', size=18, bold=True, color='1A1D24')
H2 = Font(name='Helvetica', size=13, bold=True, color='1A1D24')
LBL = Font(name='Helvetica', size=11, color='4A5160')
VAL = Font(name='Helvetica', size=11, color='1A1D24', bold=True)
ACCENT = Font(name='Helvetica', size=11, color='2563EB', bold=True)
GREEN = Font(name='Helvetica', size=11, color='16A34A', bold=True)
SMALL = Font(name='Helvetica', size=9, color='8A8F99')
MUTED = Font(name='Helvetica', size=10, color='8A8F99', italic=True)

CARD_FILL = PatternFill('solid', fgColor='FFFFFF')
INPUT_FILL = PatternFill('solid', fgColor='FFFBEB')   # subtle yellow = editable
HEADER_FILL = PatternFill('solid', fgColor='F1F3F7')
THIN = Side(border_style='thin', color='E1E4EA')
MED = Side(border_style='thin', color='C5CAD3')
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_BOX = Border(left=THIN, right=THIN, top=MED, bottom=MED)

CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)
RIGHT = Alignment(horizontal='right', vertical='center')


def build():
    wb = Workbook()
    ws = wb.active
    ws.title = 'Calculator'

    # Column widths
    widths = [18, 16, 12, 11, 11, 11, 11, 11, 11]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ---------- TITLE ----------
    ws['A1'] = 'Credit Payment Calculator'
    ws['A1'].font = H1
    ws.row_dimensions[1].height = 26

    # ---------- INPUTS BOX ----------
    ws['A3'] = 'Inputs (yellow = edit me)'
    ws['A3'].font = H2

    ws['A4'] = 'Province'
    ws['A4'].font = LBL
    ws['B4'] = 'British Columbia'
    ws['B4'].font = VAL
    ws['B4'].fill = INPUT_FILL
    ws['B4'].border = BOX

    ws['A5'] = 'Max plan length (months)'
    ws['A5'].font = LBL
    ws['B5'] = 24
    ws['B5'].font = VAL
    ws['B5'].fill = INPUT_FILL
    ws['B5'].border = BOX
    ws['B5'].number_format = '0" mo"'

    ws['A6'] = 'Down payment ($)'
    ws['A6'].font = LBL
    ws['B6'] = 0
    ws['B6'].font = VAL
    ws['B6'].fill = INPUT_FILL
    ws['B6'].border = BOX
    ws['B6'].number_format = '$#,##0.00'

    # Province dropdown
    province_names = ','.join(f'"{p[1]}"' for p in PROVINCES)
    dv_prov = DataValidation(type='list', formula1=f'"{",".join(p[1] for p in PROVINCES)}"', allow_blank=False)
    ws.add_data_validation(dv_prov)
    dv_prov.add('B4')

    # Max plan length dropdown
    dv_max = DataValidation(type='list', formula1='"6,12,18,24,45"', allow_blank=False)
    ws.add_data_validation(dv_max)
    dv_max.add('B5')

    # Tax rate (lookup)
    ws['A8'] = 'Tax rate (auto)'
    ws['A8'].font = LBL
    # VLOOKUP from Reference sheet
    ws['B8'] = '=VLOOKUP(B4,Reference!B2:C14,2,FALSE)'
    ws['B8'].font = VAL
    ws['B8'].number_format = '0.000%'

    # ---------- QUICK ESTIMATE ----------
    # Single-amount entry that instantly shows longest eligible Flexiti + MHC.
    ws['D3'] = 'Quick estimate'
    ws['D3'].font = H2

    ws['D4'] = 'Amount before tax ($)'
    ws['D4'].font = LBL
    ws['E4'] = 1000
    ws['E4'].font = VAL
    ws['E4'].fill = INPUT_FILL
    ws['E4'].border = BOX
    ws['E4'].number_format = '$#,##0.00'

    ws['D5'] = 'Total incl. tax'
    ws['D5'].font = LBL
    ws['E5'] = '=E4*(1+B8)'
    ws['E5'].font = VAL
    ws['E5'].number_format = '$#,##0.00'

    ws['D6'] = 'Credit (after down)'
    ws['D6'].font = LBL
    ws['E6'] = '=MAX(0,E5-B6)'
    ws['E6'].font = ACCENT
    ws['E6'].number_format = '$#,##0.00'

    # Longest Flexiti plan that fits both: months <= max plan length AND credit >= minAmount.
    # Build nested IF outer→inner = largest→smallest term so the LARGEST eligible
    # term wins. Iterate ascending so each step wraps the previous as a fallback,
    # making the final outermost IF the largest term.
    flexiti_sorted = sorted(FLEXITI, key=lambda x: x[0])  # asc → outermost = largest
    flex_formula = '"-"'
    for months, adminFee, minAmount in flexiti_sorted:
        cond = f'AND({months}<=$B$5,$E$6>={minAmount})'
        val = f'"{months} mo: $"&TEXT(($E$6+{adminFee})/{months},"#,##0.00")&"/mo"'
        flex_formula = f'IF({cond},{val},{flex_formula})'

    ws['D8'] = 'Longest Flexiti'
    ws['D8'].font = LBL
    ws['E8'] = f'=IF($E$6=0,"-",{flex_formula})'
    ws['E8'].font = ACCENT

    mhc_sorted = sorted(MHC, key=lambda x: x[0])  # asc → outermost = largest
    mhc_formula = '"-"'
    for months, _, annualFee, minAmount in mhc_sorted:
        years = -(-months // 12)
        annual_total = annualFee * years
        cond = f'AND({months}<=$B$5,$E$6>={minAmount})'
        val = f'"{months} mo: $"&TEXT(($E$6+{annual_total})/{months},"#,##0.00")&"/mo"'
        mhc_formula = f'IF({cond},{val},{mhc_formula})'

    ws['D9'] = 'Longest MHC'
    ws['D9'].font = LBL
    ws['E9'] = f'=IF($E$6=0,"-",{mhc_formula})'
    ws['E9'].font = GREEN

    # Widen E
    ws.column_dimensions['D'].width = 22
    ws.column_dimensions['E'].width = 22

    # ---------- ITEMS ----------
    ws['A10'] = 'Items (before tax)'
    ws['A10'].font = H2

    ws['A11'] = 'Item name'
    ws['B11'] = 'Amount ($)'
    ws['C11'] = 'Cumulative'
    for col in ('A11', 'B11', 'C11'):
        ws[col].font = LBL
        ws[col].fill = HEADER_FILL
        ws[col].border = HEADER_BOX
        ws[col].alignment = CENTER

    # 8 item rows (rows 12-19) — Item 1 pre-filled with $1000 so tables aren't
    # empty on first open. User can edit any cell.
    sample_amounts = [1000, 0, 0, 0, 0, 0, 0, 0]
    for i in range(8):
        r = 12 + i
        ws.cell(r, 1, value=f'Item {i+1}').font = VAL
        ws.cell(r, 1).fill = INPUT_FILL
        ws.cell(r, 1).border = BOX
        ws.cell(r, 2, value=sample_amounts[i]).fill = INPUT_FILL
        ws.cell(r, 2).border = BOX
        ws.cell(r, 2).number_format = '$#,##0.00'
        # Cumulative
        ws.cell(r, 3, value=f'=SUM($B$12:B{r})')
        ws.cell(r, 3).font = ACCENT
        ws.cell(r, 3).border = BOX
        ws.cell(r, 3).number_format = '$#,##0.00'

    # ---------- TOTALS ----------
    ws['A21'] = 'Totals'
    ws['A21'].font = H2

    ws['A22'] = 'Subtotal'
    ws['A22'].font = LBL
    ws['B22'] = '=SUM(B12:B19)'
    ws['B22'].font = VAL
    ws['B22'].number_format = '$#,##0.00'

    ws['A23'] = 'Tax'
    ws['A23'].font = LBL
    ws['B23'] = '=B22*B8'
    ws['B23'].font = VAL
    ws['B23'].number_format = '$#,##0.00'

    ws['A24'] = 'Total (incl. tax)'
    ws['A24'].font = LBL
    ws['B24'] = '=B22+B23'
    ws['B24'].font = ACCENT
    ws['B24'].number_format = '$#,##0.00'

    ws['A25'] = 'Down payment'
    ws['A25'].font = LBL
    ws['B25'] = '=B6'
    ws['B25'].font = VAL
    ws['B25'].number_format = '$#,##0.00'

    ws['A26'] = '20% suggested'
    ws['A26'].font = SMALL
    ws['B26'] = '=B24*0.2'
    ws['B26'].font = SMALL
    ws['B26'].number_format = '$#,##0.00'

    ws['A27'] = 'Credit amount'
    ws['A27'].font = LBL
    ws['B27'] = '=MAX(0,B24-B25)'
    ws['B27'].font = ACCENT
    ws['B27'].number_format = '$#,##0.00'

    # ---------- FLEXITI TABLE ----------
    ws['A29'] = 'Flexiti — monthly payment by plan'
    ws['A29'].font = H2
    ws['A30'] = 'Cumulative incl. tax less down'
    ws['A30'].font = LBL
    ws['A30'].alignment = LEFT

    # Term column headers (B=3mo, C=6mo, D=12mo, E=18mo, F=24mo, G=45mo)
    flexiti_cols = []
    for j, (months, adminFee, minAmount) in enumerate(FLEXITI):
        col_letter = get_column_letter(2 + j)
        flexiti_cols.append((col_letter, months, adminFee, minAmount))
        ws.cell(30, 2 + j, value=f'{months} mo\n+${adminFee:.2f} fee').font = LBL
        ws.cell(30, 2 + j).fill = HEADER_FILL
        ws.cell(30, 2 + j).border = HEADER_BOX
        ws.cell(30, 2 + j).alignment = CENTER

    ws.row_dimensions[30].height = 30

    # Per-row Flexiti monthlies (rows 31-38 mirror items 12-19)
    for i in range(8):
        item_r = 12 + i
        out_r = 31 + i
        # Credit for this cumulative row (incl. tax less down payment, only if >0)
        ws.cell(out_r, 1, value=f'=IF(C{item_r}=0,"",MAX(0,C{item_r}*(1+$B$8)-$B$25))')
        ws.cell(out_r, 1).font = VAL
        ws.cell(out_r, 1).border = BOX
        ws.cell(out_r, 1).number_format = '$#,##0.00'

        for j, (col_letter, months, adminFee, minAmount) in enumerate(flexiti_cols):
            credit_ref = f'A{out_r}'
            # Only show if: cumulative populated AND term within max-plan-length AND credit >= minAmount
            cell = ws.cell(out_r, 2 + j)
            cell.value = (
                f'=IF(C{item_r}=0,"",'
                f'IF({months}>$B$5,"-",'
                f'IF({credit_ref}<{minAmount},"Min ${minAmount:.0f}",'
                f'({credit_ref}+{adminFee})/{months})))'
            )
            cell.font = ACCENT
            cell.border = BOX
            cell.number_format = '"$"#,##0.00"/mo"'
            cell.alignment = RIGHT

    # ---------- MHC TABLE ----------
    ws['A40'] = 'MHC — monthly payment by plan'
    ws['A40'].font = H2
    ws['A41'] = 'Cumulative incl. tax less down'
    ws['A41'].font = LBL

    mhc_cols = []
    for j, (months, adminFee, annualFee, minAmount) in enumerate(MHC):
        col_letter = get_column_letter(2 + j)
        mhc_cols.append((col_letter, months, adminFee, annualFee, minAmount))
        ws.cell(41, 2 + j, value=f'{months} mo\n${annualFee}/yr fee').font = LBL
        ws.cell(41, 2 + j).fill = HEADER_FILL
        ws.cell(41, 2 + j).border = HEADER_BOX
        ws.cell(41, 2 + j).alignment = CENTER

    ws.row_dimensions[41].height = 30

    for i in range(8):
        item_r = 12 + i
        out_r = 42 + i
        ws.cell(out_r, 1, value=f'=IF(C{item_r}=0,"",MAX(0,C{item_r}*(1+$B$8)-$B$25))')
        ws.cell(out_r, 1).font = VAL
        ws.cell(out_r, 1).border = BOX
        ws.cell(out_r, 1).number_format = '$#,##0.00'

        for j, (col_letter, months, adminFee, annualFee, minAmount) in enumerate(mhc_cols):
            credit_ref = f'A{out_r}'
            years = -(-months // 12)  # ceil division
            annual_total = annualFee * years
            cell = ws.cell(out_r, 2 + j)
            cell.value = (
                f'=IF(C{item_r}=0,"",'
                f'IF({months}>$B$5,"-",'
                f'IF({credit_ref}<{minAmount},"Min ${minAmount:.0f}",'
                f'({credit_ref}+{annual_total})/{months})))'
            )
            cell.font = GREEN
            cell.border = BOX
            cell.number_format = '"$"#,##0.00"/mo"'
            cell.alignment = RIGHT

    # ---------- INSTRUCTIONS ----------
    ws['A51'] = 'How it works'
    ws['A51'].font = H2
    notes = [
        '1. Pick your province in B4 — tax rate auto-fills.',
        '2. Set max plan length in B5 (6/12/18/24/45 months) — caps which terms appear.',
        '3. Enter line items in A12:B19. The "Cumulative" column shows running totals.',
        '4. Optionally enter a down payment in B6 (or use the 20% suggestion in B26).',
        '5. Flexiti and MHC tables compute monthly payment for each cumulative subtotal.',
        '6. "Min $X" = below tier minimum.  "—" = exceeds your max plan length.',
    ]
    for i, line in enumerate(notes):
        ws.cell(52 + i, 1, value=line).font = MUTED
        ws.merge_cells(start_row=52 + i, start_column=1, end_row=52 + i, end_column=8)

    # ---------- REFERENCE SHEET ----------
    ref = wb.create_sheet('Reference')
    ref.column_dimensions['A'].width = 6
    ref.column_dimensions['B'].width = 28
    ref.column_dimensions['C'].width = 12
    ref['A1'] = 'Code'; ref['B1'] = 'Province'; ref['C1'] = 'Tax rate'
    for c in ('A1', 'B1', 'C1'):
        ref[c].font = Font(bold=True)
    for i, (code, name, rate) in enumerate(PROVINCES):
        ref.cell(2 + i, 1, value=code)
        ref.cell(2 + i, 2, value=name)
        ref.cell(2 + i, 3, value=rate).number_format = '0.000%'

    ref['E1'] = 'Flexiti tier'; ref['F1'] = 'Months'; ref['G1'] = 'Admin fee'; ref['H1'] = 'Min credit'
    for c in ('E1', 'F1', 'G1', 'H1'):
        ref[c].font = Font(bold=True)
    for i, (m, fee, min_) in enumerate(FLEXITI):
        ref.cell(2 + i, 5, value=f'F-{m}')
        ref.cell(2 + i, 6, value=m)
        ref.cell(2 + i, 7, value=fee).number_format = '$#,##0.00'
        ref.cell(2 + i, 8, value=min_).number_format = '$#,##0'

    ref['J1'] = 'MHC tier'; ref['K1'] = 'Months'; ref['L1'] = 'Annual fee'; ref['M1'] = 'Min credit'
    for c in ('J1', 'K1', 'L1', 'M1'):
        ref[c].font = Font(bold=True)
    for i, (m, _, ann, min_) in enumerate(MHC):
        ref.cell(2 + i, 10, value=f'M-{m}')
        ref.cell(2 + i, 11, value=m)
        ref.cell(2 + i, 12, value=ann).number_format = '$#,##0'
        ref.cell(2 + i, 13, value=min_).number_format = '$#,##0'

    out = '/home/bach/projects/mh/standalone/credit_calculator.xlsx'
    wb.save(out)
    return out


if __name__ == '__main__':
    out = build()
    print(f'Wrote {out}')
